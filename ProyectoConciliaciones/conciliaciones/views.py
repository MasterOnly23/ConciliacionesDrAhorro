from typing import Any
import os
from django.shortcuts import render
from django.views import View
from django.http import JsonResponse
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
import pandas as pd
from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO
from django.core.exceptions import ValidationError
from conciliaciones.models import (
    FileHeaders,
    Extractos,
    Mayor,
    Conciliacion,
    NoConciliado,
)
from datetime import datetime
import numpy as np
import logging
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.utils import timezone
from django.core import serializers
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from decouple import config
from openpyxl.styles import Font, Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)
logger.setLevel(logging.WARNING)
handler = logging.handlers.RotatingFileHandler(
    settings.LOGS_PATH + "conciliaciones.log", maxBytes=5 * 1024 * 1024, backupCount=5
)
logger.addHandler(handler)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
handler.setFormatter(formatter)

# Create your views here.


@method_decorator(csrf_exempt, name="dispatch")
class ConciliacionesView(View):
    FILE_SIZE_LIMIT = 5 * 1024 * 1024  # 5MB

    def post(self, request):
        self.excel_file = request.FILES["file"]
        file_name = self.excel_file.name
        print(file_name)

        # Verifica la extensión del archivo
        if not self.excel_file.name.endswith(
            ".xls"
        ) and not self.excel_file.name.endswith(".xlsx"):
            return JsonResponse(
                {"error": "Invalid file type. Only .xls and .xlsx are accepted."},
                status=400,
            )

        # Verifica el tamaño del archivo (5MB en este ejemplo)
        if self.excel_file.size > self.FILE_SIZE_LIMIT:
            return JsonResponse(
                {"error": "File is too large. Maximum file size is 5MB."}
            )
        if not file_name.lower().startswith(config("FILE_NAME_PREFIX")):
            return JsonResponse({"error": "Nombre de archivo no Valido."}, status=400)

        newFileName = timezone.now().strftime("%Y-%m-%d") + "-" + file_name
        path = default_storage.save(
            "files/" + newFileName, ContentFile(self.excel_file.read())
        )

        data_extractos = pd.read_excel(
            self.excel_file, sheet_name="EXTRACTO", header=None
        )
        data_mayor = pd.read_excel(self.excel_file, sheet_name="MAYOR", header=None)

        self.bank_name = data_extractos.iat[0, 0].upper()  # Accede a la celda A1
        self.period = data_extractos.iat[0, 1]

        try:
            file_header = FileHeaders.objects.create(
                name=file_name,
                bank_name=self.bank_name,
                periodo=self.period,
            )
            print(file_header.id)
            self.create_extractos(data_extractos, file_header)
            self.create_mayor(data_mayor, file_header)
            self.conciliacion(self.bank_name, file_header)
            return JsonResponse(
                {
                    "message": "Excel file has been processed.",
                    "bank_name": self.bank_name,
                    "periodo": self.period,
                    "file_header_id" : file_header.id
                }
            )
        except Exception as e:
            return JsonResponse({"error": str(e)})

    def create_extractos(self, data_extractos, file_header):
        data_table_estractos = data_extractos.iloc[3:, :5]  # leemos la tabla
        extractos_data = []
        for index, row in data_table_estractos.iterrows():
            try:
                if isinstance(row[0], datetime):
                    fecha = row[0].strftime("%Y-%m-%d")
                elif isinstance(row[0], str):
                    fecha = datetime.strptime(row[0], "%d/%m/%Y").strftime("%Y-%m-%d")
                else:
                    fecha = None
                extractos_data.append(
                    Extractos(
                        file_header=file_header,
                        fecha=fecha,
                        descripcion=row[1],
                        comprobante=row[2],
                        monto=row[3],
                        codigo=row[4],
                    )
                )
            except Exception as e:
                logger.error(str(e))
        Extractos.objects.bulk_create(extractos_data)

    def create_mayor(self, data_mayor, file_header):
        data_table_mayor = data_mayor.iloc[3:, :6]
        mayor_data = []
        for index, row in data_table_mayor.iterrows():
            try:
                if isinstance(row[0], datetime):
                    fecha = row[0].strftime("%Y-%m-%d")
                elif isinstance(row[0], str):
                    fecha = datetime.strptime(row[0], "%d/%m/%Y").strftime("%Y-%m-%d")
                else:
                    fecha = None
                mayor_data.append(
                    Mayor(
                        file_header=file_header,
                        fecha=fecha,
                        descripcion=row[1],
                        monto=row[2] if not np.isnan(row[2]) else None,
                        codigo=row[3],
                    )
                )
            except Exception as e:
                logger.error(str(e))
        Mayor.objects.bulk_create(mayor_data)

    def conciliacion(self, bank_name, file_header):
        try:
            extractos = Extractos.objects.filter(file_header=file_header)
            mayor = Mayor.objects.filter(file_header=file_header)
            no_conciliados = []
            # file_header = FileHeaders.objects.filter(bank_name=bank_name, periodo=self.period).last()

            # Sumatoria 6666
            extractos_6666 = [e for e in extractos if e.codigo == '6666']
            sum_extractos_6666 = sum(e.monto for e in extractos_6666)

            #Sumatoria 1111
            extractos_1111 = [e for e in extractos if e.codigo == '1111']
            sum_extractos_1111 = sum(e.monto for e in extractos_1111)

            try:
                for e in extractos:
                    conciliado = False
                    if e.codigo != '6666' and e.codigo != '1111':
                        for m in mayor:
                            if (
                                # e.fecha == m.fecha
                                e.codigo == m.codigo
                                and e.monto == m.monto
                            ):
                                Conciliacion.objects.create(
                                    file_header=file_header,
                                    extracto=e,
                                    mayor=m,
                                )
                                conciliado = True
                                break
                        if not conciliado:
                            no_conciliado = NoConciliado.objects.create(
                                file_header=file_header,
                                extracto_fecha=e.fecha,
                                extracto_descripcion=e.descripcion,
                                extracto_monto=e.monto,
                                extracto_codigo=e.codigo,
                            )
                            no_conciliados.append(no_conciliado)

                conciliado_6666 = False
                conciliado_1111 = False
                for m in mayor:
                    if m.codigo == '6666':
                        if sum_extractos_6666 == m.monto:
                            Conciliacion.objects.create(
                                file_header=file_header,
                                extracto=extractos_6666[0],
                                mayor=m,
                            )
                            conciliado_6666 = True
                            # break
                        else:
                            no_conciliado = NoConciliado.objects.create(
                                file_header=file_header,
                                mayor_fecha=m.fecha,
                                mayor_descripcion=m.descripcion,
                                mayor_monto=m.monto,
                                mayor_codigo=m.codigo,
                            )
                            no_conciliados.append(no_conciliado)
                    #     break
                    # if not conciliado_6666:
                    #     no_conciliado = NoConciliado.objects.create(
                    #         file_header=file_header,
                    #         extracto_fecha=extractos_6666[0].fecha,
                    #         extracto_descripcion='Gastos Bancarios',
                    #         extracto_monto=sum_extractos_6666,
                    #         extracto_codigo='6666',
                    #     )
                    #     no_conciliados.append(no_conciliado)
                    
                    elif m.codigo == '1111':
                        if sum_extractos_1111 == m.monto:
                            Conciliacion.objects.create(
                                file_header=file_header,
                                extracto=extractos_1111[0],
                                mayor=m,
                            )
                            conciliado_1111 = True
                            # break
                        else:
                            no_conciliado = NoConciliado.objects.create(
                                file_header=file_header,
                                mayor_fecha=m.fecha,
                                mayor_descripcion=m.descripcion,
                                mayor_monto=m.monto,
                                mayor_codigo=m.codigo,
                            )
                            no_conciliados.append(no_conciliado)
                    # else:
                    #     no_conciliado = NoConciliado.objects.create(
                    #         file_header=file_header,
                    #         mayor_fecha=m.fecha,
                    #         mayor_descripcion=m.descripcion,
                    #         mayor_monto=m.monto,
                    #         mayor_codigo=m.codigo,
                    #     )
                    #     no_conciliados.append(no_conciliado)

                if not conciliado_6666 and extractos_6666:
                    no_conciliado = NoConciliado.objects.create(
                        file_header=file_header,
                        extracto_fecha=extractos_6666[0].fecha,
                        extracto_descripcion='Gastos Bancarios',
                        extracto_monto=sum_extractos_6666,
                        extracto_codigo='6666',
                    )
                    no_conciliados.append(no_conciliado)
                    
                
                if not conciliado_1111 and extractos_1111:
                    no_conciliado = NoConciliado.objects.create(
                        file_header=file_header,
                        extracto_fecha=extractos_1111[0].fecha,
                        extracto_descripcion='Valores depositados en efectivo',
                        extracto_monto=sum_extractos_1111,
                        extracto_codigo='1111',
                    )
                    no_conciliados.append(no_conciliado)

            except ValidationError as e:
                print(e)
                logger.error(str(e))
                return JsonResponse({"error": str(e)})

            try:
                for m in mayor:
                    conciliado = False
                    if m.codigo != '6666' and m.codigo != '1111':
                        for e in extractos:
                            if (
                                # m.fecha == e.fecha
                                m.codigo == e.codigo
                                and m.monto == e.monto
                            ):
                                Conciliacion.objects.create(
                                    file_header=file_header,
                                    extracto=e,
                                    mayor=m,
                                )
                                conciliado = True
                                break
                        if not conciliado:
                            no_conciliado, created = NoConciliado.objects.get_or_create(
                                file_header=file_header,
                                mayor_fecha=m.fecha,
                                mayor_descripcion=m.descripcion,
                                mayor_monto=m.monto,
                                mayor_codigo=m.codigo,
                            )
                            if created:
                                no_conciliados.append(no_conciliado)
                

                    
            except ValidationError as e:
                print(e)
                logger.error(str(e))
                return JsonResponse({"error": str(e)})

            return JsonResponse(
                {
                    "message": "Excel file has been processed.",
                    # "no_conciliados": [
                    #     {
                    #         "extracto": {
                    #             "fecha": nc.extracto_fecha,
                    #             "descripcion": nc.extracto_descripcion,
                    #             "monto": nc.extracto_monto,
                    #         }if nc.extracto_fecha is not None else None,
                    #         "mayor": {
                    #             "fecha": nc.mayor_fecha,
                    #             "descripcion": nc.mayor_descripcion,
                    #             "monto": nc.mayor_monto,
                    #         }if nc.mayor_fecha is not None else None,
                    #     } for nc in no_conciliados
                    # ],
                }
            )

        except Exception as e:
            print(e)
            logger.error(str(e))
            return JsonResponse({"error": str(e)})

    def get(self, request, *args, **kwargs):
        bankName = request.GET.get("bankName")
        period = request.GET.get("period")
        action = request.GET.get("action")
        monto = request.GET.get("monto")
        limit = request.GET.get("limit", 10)
        file_header_id = request.GET.get("fileId")
        if action == "diferencias":
            if bankName and period:
                try:
                    file_header = FileHeaders.objects.filter(
                        bank_name=bankName, periodo=period
                    ).last()
                    print(file_header.id)
                    extractos = Extractos.objects.filter(file_header=file_header.id)
                    mayores = Mayor.objects.filter(file_header=file_header.id)
                    conciliaciones = Conciliacion.objects.filter(
                        file_header=file_header.id
                    )
                    no_conciliados = NoConciliado.objects.filter(
                        file_header=file_header.id
                    )
                    return JsonResponse(
                        {
                            "extractos": [
                                {
                                    "fecha": e.fecha,
                                    "descripcion": e.descripcion,
                                    "comprobante": e.comprobante,
                                    "monto": e.monto,
                                    "codigo": e.codigo,
                                }
                                for e in extractos
                            ],
                            "mayores": [
                                {
                                    "fecha": m.fecha,
                                    "descripcion": m.descripcion,
                                    "monto": m.monto,
                                    "codigo": m.codigo,
                                }
                                for m in mayores
                            ],
                            "conciliaciones": [
                                {
                                    "extracto": {
                                        "fecha": c.extracto.fecha,
                                        "descripcion": c.extracto.descripcion,
                                        "comprobante": c.extracto.comprobante,
                                        "monto": c.extracto.monto,
                                        "codigo": c.extracto.codigo,
                                    },
                                    "mayor": {
                                        "fecha": c.mayor.fecha,
                                        "descripcion": c.mayor.descripcion,
                                        "monto": c.mayor.monto,
                                        "codigo": c.mayor.codigo,
                                    },
                                }
                                for c in conciliaciones
                            ],
                            "no_conciliados": [
                                {
                                    "extracto": (
                                        {
                                            "fecha": nc.extracto_fecha,
                                            "descripcion": nc.extracto_descripcion,
                                            "monto": nc.extracto_monto,
                                            "codigo": nc.extracto_codigo,
                                        }
                                        if nc.extracto_fecha is not None
                                        else None
                                    ),
                                    "mayor": (
                                        {
                                            "fecha": nc.mayor_fecha,
                                            "descripcion": nc.mayor_descripcion,
                                            "monto": nc.mayor_monto,
                                            "codigo": nc.mayor_codigo,
                                        }
                                        if nc.mayor_fecha is not None
                                        else None
                                    ),
                                }
                                for nc in no_conciliados
                            ],
                            "file_header_id": file_header.id,
                        }
                    )
                except FileHeaders.DoesNotExist:
                    return JsonResponse(
                        {"error": "No file found for the given bank and period."}
                    )
        elif action == "historial":
            filter_args = {}
            if bankName:
                filter_args["bank_name"] = bankName
            if period:
                filter_args["periodo"] = period
            if monto:
                filter_args["monto"] = monto
            if filter_args:
                file_headers = FileHeaders.objects.filter(**filter_args)
            else:
                file_headers = FileHeaders.objects.all()

            if limit:
                file_headers = file_headers[: int(limit)]

            response_data = []
            for file_header in file_headers:
                no_conciliados = NoConciliado.objects.filter(file_header=file_header.id)
                no_conciliados_json = serializers.serialize("json", no_conciliados)
                response_data.append(
                    {
                        "file_header": serializers.serialize("json", [file_header]),
                        "file_id": file_header.id,
                        "no_conciliados": no_conciliados_json,
                    }
                )
            return JsonResponse(response_data, safe=False)

        elif action == "export":
            if not file_header_id:
                return JsonResponse(
                    {"error": "file_header_id es requerido."}, status=400
                )
            file_header = FileHeaders.objects.get(id=file_header_id)
            no_conciliados = NoConciliado.objects.filter(file_header=file_header.id)
            wb = Workbook()
            ws_extracto = wb.create_sheet("Extracto")
            ws_mayor = wb.create_sheet("Mayor")
            header_font = Font(bold=True)
            center_alignment = Alignment(horizontal="center")
            extracto_fill = PatternFill(
                start_color="95B3D7", end_color="95B3D7", fill_type="solid"
            )
            mayor_fill = PatternFill(
                start_color="C4D79B", end_color="C4D79B", fill_type="solid"
            )
            headers_extracto = ["Extracto Fecha", "Extracto Descripcion", "Extracto Monto", "Extracto Codigo"]
            headers_mayor = ["Mayor Fecha", "Mayor Descripcion", "Mayor Monto", "Mayor Codigo"]
            for i, header in enumerate(headers_extracto, start=1):
                cell = ws_extracto.cell(row=1, column=i, value=header)
                cell.font = header_font
                cell.alignment = center_alignment
                cell.fill = extracto_fill
            for i, header in enumerate(headers_mayor, start=1):
                cell = ws_mayor.cell(row=1, column=i, value=header)
                cell.font = header_font
                cell.alignment = center_alignment
                cell.fill = mayor_fill
            for i, no_conciliado in enumerate(no_conciliados, start=2):
                ws_extracto.cell(row=i, column=1, value=no_conciliado.extracto_fecha)
                ws_extracto.cell(row=i, column=2, value=no_conciliado.extracto_descripcion)
                ws_extracto.cell(row=i, column=3, value=no_conciliado.extracto_monto)
                ws_extracto.cell(row=i, column=4, value=no_conciliado.extracto_codigo)
                ws_mayor.cell(row=i, column=1, value=no_conciliado.mayor_fecha)
                ws_mayor.cell(row=i, column=2, value=no_conciliado.mayor_descripcion)
                ws_mayor.cell(row=i, column=3, value=no_conciliado.mayor_monto)
                ws_mayor.cell(row=i, column=4, value=no_conciliado.mayor_codigo)
            wb.remove(wb["Sheet"])

            for sheet in wb.sheetnames:
                for column in wb[sheet].columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2  # Ajuste para la medida de Excel
                    wb[sheet].column_dimensions[
                        column[0].column_letter
                    ].width = adjusted_width

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = "attachment; filename=NoConciliados.xlsx"
            wb.save(response)
            return response

        else:
            return JsonResponse({"error": "Invalid action."})


def DownloadPlantilla(request):
    bankName = request.GET.get("bankName")
    period = request.GET.get("period")
    try:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        ws_extracto = wb.create_sheet("EXTRACTO")
        ws_extracto["A1"] = bankName
        ws_extracto["A1"].font = Font(name="Calibri", size=20, bold=True)
        ws_extracto["A1"].alignment = Alignment(horizontal="left")

        ws_extracto["B1"] = period
        ws_extracto["B1"].font = Font(name="Calibri", size=20, bold=True)
        ws_extracto["B1"].alignment = Alignment(horizontal="left")

        header_font = Font(name="Calibri", bold=True)

        headers = ["Fecha", "Descripcion", "Comprobante", "Importe", "Codigo"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_extracto.cell(row=3, column=col_num, value=header)
            cell.font = header_font

        for column in ws_extracto.columns:
            max_length = 0
            column = get_column_letter(
                column[0].column
            )  # Obtener la letra de la columna
            for cell in ws_extracto[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            ws_extracto.column_dimensions[column].width = adjusted_width

        # ajusta size A1 B1
        ws_extracto.column_dimensions[get_column_letter(1)].width = 30
        ws_extracto.column_dimensions[get_column_letter(2)].width = 30

        # MAYOR
        ws_mayor = wb.create_sheet("MAYOR")
        ws_mayor["A1"] = bankName
        ws_mayor["A1"].font = Font(name="Calibri", size=20, bold=True)
        ws_mayor["A1"].alignment = Alignment(horizontal="left")

        ws_mayor["B1"] = period
        ws_mayor["B1"].font = Font(name="Calibri", size=20, bold=True)
        ws_mayor["B1"].alignment = Alignment(horizontal="left")

        headers = ["Fecha", "Descripcion", "Importe", "Codigo"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_mayor.cell(row=3, column=col_num, value=header)
            cell.font = header_font

        for column in ws_mayor.columns:
            max_length = 0
            column = get_column_letter(
                column[0].column
            )  # Obtener la letra de la columna
            for cell in ws_mayor[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            ws_mayor.column_dimensions[column].width = adjusted_width

        # ajusta size A1 B1
        ws_mayor.column_dimensions[get_column_letter(1)].width = 30
        ws_mayor.column_dimensions[get_column_letter(2)].width = 30

        # BytesIO guarda el objeto en memoria y no en el disco. se limpia al terminar la funcion
        output = BytesIO()

        wb.save(output)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename=Conciliaciones {bankName}.xlsx"
        )
        return response
    except Exception as e:
        logger.error(str(e))
        return JsonResponse({"error": str(e)})
